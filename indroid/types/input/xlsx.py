"""Get data from Excel 2007-2010 file, xlsx. Find the header names heuristically, ignoring extra rows at the beginning.
For reading Excel 1997-2003, use the xls_-module.
TBD: enable 'raw' reading, without fieldnames. This must be enabled for all sources!"""
import os, sys
if os.getcwd() in sys.path:
    sys.path.remove(os.getcwd())
import openpyxl as xls
from datetime import datetime
from indroid.input.general import *

class Row(Row):
    def __init__(self, row, fieldnames=None):
        super(Row, self).__init__()
        for fieldname, cell in zip(fieldnames, row):
            self[fieldname] = cell.internal_value

class Table(Table):
    def __init__(self, sheet, max_rows=None, guess_fieldnames=True, fieldnames=[], raw=False):
        assert isinstance(sheet, xls.worksheet.Worksheet)
        self.sheet = sheet
        self.name = sheet.title
        self.max_rows = max_rows
        # Find the row with fieldnames, assuming it has fieldnames:
        
        i = -1
        row = []
        for i, row in enumerate(sheet.iter_rows()):
            if all([cell.internal_value for cell in row]):
                break
        self.fieldnames = [unicode(cell.internal_value) for cell in row]
        self.row_offset = i
    def __iter__(self):
        for i, row in enumerate(self.sheet.iter_rows()):
            if i <= self.row_offset:
                continue
            if self.max_rows is not None and i > self.row_offset + self.max_rows:
                break
            yield Row(row, self.fieldnames)

class Source(Source):
    def __init__(self, filename, max_rows=None, guess_rownames=True):
        "The filename is stored as self.name and the path as self.path."
        self.path, name = os.path.split(filename)
        super(Source, self).__init__(name, max_rows=max_rows)
        wb = xls.load_workbook(filename, use_iterators=True)
        for sheet in wb.worksheets:
            self._tables.append(Table(sheet, max_rows, guess_rownames))

def test():
    source = Source(r"X:\data\source\kpn\swol_marketing\odin\2012-09-26\extract\odin_dump-packages.xlsx")
    rows = []
    for table in source:
        for row in table:
            rows.append(row)
    print len(rows)
    for i in (0, 285, 284, 569):
        print rows[i]
    print rows[0] == rows[285]

if __name__ == '__main__':
    test()